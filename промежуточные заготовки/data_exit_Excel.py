# Вывод расчетных данных в файл Excel

import openpyxl
from DPAS import *

book = openpyxl.Workbook()

sheet = book.active

sheet['A1'] = 'Пролетные строения'
sheet['A2'] = '№ п.п.'
sheet['B2'] = '№ в ЛСР '
sheet['C2'] = 'Наименование работ'
sheet['D2'] = 'Ед. изм.'
sheet['E2'] = 'Кол-во'

# sheet['F2'] = 'Кол.'
# sheet['A3'] = 'Опоры'
# sheet['B3'] = 'Устои'
# sheet['B27'] = 'Промежуточные опоры'
# sheet['A42'] = 'Пролетное строение'
# sheet['A54'] = 'Мостовое полотно'
# sheet['A63'] = 'Сопряжение'

sheet['C4'].value = 'Установка опорных частей РОЧ и установка клиновидных подкладок'
sheet['E4'].value = nrochust + nrochust
sheet['C5'].value = 'Монтаж балок двумя стреловыми кранами'
sheet['E5'].value = nbalok
sheet['C10'].value = 'Поперечная передвижка балок'
sheet['E10'].value = nbalok-2
sheet['C11'].value = 'Устройство средних (продольных) монолитных участков объединения балок с армированием'
sheet['E11'].value = Vmonuch
sheet['C16'].value = 'Устройство монолитных участков соединительной плиты с армированием'
sheet['E16'].value = Vplitomon
sheet['C22'].value = 'Устройство монолитных участков деформационного шва с армированием'
sheet['E22'].value = Vplitomon
sheet['C24'].value = 'Устройство монолитных карнизных блоков с армированием'
sheet['E24'].value = Vplitomon
sheet['C26'].value = 'Обмазочная гидроизоляция ростверков и тела опор'

sheet['C27'].value = 'Устройство вертикальных скважин'
sheet['C28'].value = 'Буронабивные сваи'
sheet['C35'].value = 'Покарска трубы 1020х10 мм'
sheet['C36'].value = 'Ригель'
sheet['C40'].value = 'Подферменные площадки'

sheet['C42'].value = 'Опорные части'
sheet['C43'].value = 'Балка '
sheet['C44'].value = 'Балка '
sheet['C45'].value = 'Монолитные участки '
sheet['C52'].value = 'Карнизный блок '

sheet['C54'].value = 'Устройство деформационных швов '
sheet['C56'].value = 'Напыляемая гидроизоляция 3 мм '
sheet['C57'].value = 'Выравнивающий слой '
sheet['C58'].value = 'Устройство водоотвода'
sheet['C59'].value = 'Покрытие проезжей части'
sheet['C60'].value = 'Покрытие служебных проходов'
sheet['C61'].value = 'Перильное ограждение'
sheet['C62'].value = 'Одностороннее барьреное ограждение'

sheet['C63'].value = 'Отсыпка конусов из дренирующего грунта'
sheet['C64'].value = 'Щебеночная подушка под лежень'
sheet['C65'].value = 'Щебеночная подготовка толщиной 100 мм'
sheet['C66'].value = 'Проливка щебня цементным раствором'
sheet['C67'].value = 'Переходные плиты и лежни'
sheet['C70'].value = 'Напыляемая гидроизоляция переходных плит'
sheet['C71'].value = 'Монолитный упор'
sheet['C72'].value = 'Укрепление конусов'
sheet['C74'].value = 'Покрытие проезжей части'
sheet['C75'].value = 'Покрытие служебных проходов'
sheet['C76'].value = 'Установка водоотводных телескопических лотков'

# sheet[1][0].value = 'Объемы основных работ'
# sheet[2][1].value = 'Сооружение'
# # sheet.cell(row=2, column=1).value = 'Сооружение'
# sheet[3][5].value = nbalok
# sheet.cell(row=4, column=4).value = 'Количество балок в пролетном строении'
# sheet[4][5].value = dlinaproleta
# sheet.cell(row=5,column=4).value='Длина пролетного строения'

book.save("data_exit.xlsx")
book.close()
